import openpyxl
from typing import List
import tkinter as tk
#TODO: add comments


class Inventory():
    def __init__(self, inv_name=None, inv_id=None, inv_quantity=None, min_inv_quantity=None, col_num=None):
        self.__inv_name = inv_name
        self.__inv_id = inv_id
        self.__inv_quantity = inv_quantity
        self.__min_inv_quantity = min_inv_quantity
        self.__colum_num = col_num
    

    @property
    def name(self):
        return self.__inv_name
        
    @property
    def id(self):
        return self.__inv_id
    
    @property
    def quantity(self):
        return self.__inv_quantity
    
    @property
    def min_quantity(self):
        return self.__min_inv_quantity

    @property
    def col(self):
        return self.__colum_num


class DBManager():
    
    def __init__(self, inv_id=None, num=None):
        self.__wb_name = "main_DB.xlsx"
        self.__ws_name = "Main Date"

        self.__inputed_inv_id = inv_id
        self.__inupted_inv_cnt = num

        self.wb = openpyxl.load_workbook(self.__wb_name)
        self.ws = self.wb[self.__ws_name]

        self.__all_inv_list = []
        self.__inv_name = ""
        self.__specified_inv_list = []
        
        self.__all_inv_items_cnt = 0
        self.__inv_quantity = 0
        self.__min_inv_quantity = 0
        self.__col_num = 0
        self.remaind_num = None           # to clarify '0' or 'not defined'
        self.is_handoverable_flag = None  # to clarify 'T/F' or 'not defined'

        self.__name_idx = 2
        self.__quantity_idx = 3
        self.__min_quantity_idx = 4


    def create_inventory(self, inv_id):
        self.__inputed_inv_id = inv_id
        self.set_params()
        return Inventory(self.__inv_name, self.__inputed_inv_id, self.__inv_quantity, self.__min_inv_quantity, self.__col_num)

    def update_db(self, inv_list: List[Inventory], button_states: List[tk.Button], spin_box_list: List[tk.Spinbox]):

        for inv, state, req_num in zip(inv_list, button_states, spin_box_list):
            cell = self.ws.cell(row=inv.col, column=6)
            print(cell.value)

            if state == "Take":
                cell.value = int(inv.quantity) - int(req_num)
            
            elif state == "Return":
                cell.value = int(inv.quantity) + int(req_num)

            else:
                 print("undefined situaion")
            
            print("{} so, remian items are {}".format(state, cell.value))

        self.close_save()


    def set_params(self):
        self.__all_inv_list = self._set_all_inv_lists()

        self.__all_inv_items_cnt = self._set_all_inv_items_cnt()
        self.__specified_inv_list = self._set_specified_inv_list()
        
        #if inputed inventory name does not exist, specified_inv_list=[] so,
        if not self.__specified_inv_list:
            self.__inv_name = None
            self.__inv_quantity = 0
            self.__min_inv_quantity = 0
            self.__col_num = None
        
        else:
            self.__inv_name = self._set_inv_name()
            self.__inv_quantity = self._set_inv_quantity()
            self.__min_inv_quantity = self._set_min_inv_quantity() 
            self.__col_num = self._set_col_num()

    def is_empty(self, cell):
        return cell.value is None or not str(cell.value).strip()

    def _set_all_inv_lists(self):
        all_inv_list = []
        for row in self.ws.iter_rows(min_row=1):
            if all(self.is_empty(c) for c in row):
                break
            values = []
            for col in row:
                values.append(col.value)
            all_inv_list.append(values)
        
        return all_inv_list

    def _set_all_inv_items_cnt(self):
        return len(self.__all_inv_list)

    def _set_specified_inv_list(self):
        specified_inv_list = []
        for l in range(self.__all_inv_items_cnt):
            if self.__all_inv_list[l][0] == self.__inputed_inv_id:
                specified_inv_list = [l+1, self.__all_inv_list[l][0], 
                                        self.__all_inv_list[l][1], 
                                        self.__all_inv_list[l][5], 
                                        self.__all_inv_list[l][6]]

        return specified_inv_list
    
    def _set_col_num(self):
        return self.__specified_inv_list[0]

    def _set_inv_name(self):
        return self.__specified_inv_list[self.__name_idx]

    def _set_inv_quantity(self):
        return self.__specified_inv_list[self.__quantity_idx]
    
    def _set_min_inv_quantity(self):
        return self.__specified_inv_list[self.__min_quantity_idx]

    def update_inv_test(self):
        # THIS IS JUST TEST
        #TODO Delete this function
        self.is_handoverable_flag = self.is_inv_handoverable()

        if self.is_handoverable_flag:
            self.remaind_num = self.calc_remaind_num()
            print("OK", self.remaind_num)

        else:
            print("can't handover {}. there are only {} inventories.".format(self.__inv_name, self.__inv_quantity))

        if self.is_handoverable_flag and not self.remaind_num:
            print("{} is complatery out of stock. ".format(self.__inv_name))

        elif self.is_handoverable_flag and self.remaind_num <= self.__min_inv_quantity:
            print("{} will be out of stock soon. There are only {}.".format(self.__inv_name, self.remaind_num))

    def calc_remaind_num(self):
        if self.__inv_quantity == None:
            raise ValueError("quantity of inventory does not defined")

        remaind_num = int(self.__inv_quantity) - int(self.__inupted_inv_cnt)
        
        return remaind_num

    def is_inv_handoverable(self):
        remaind_num = self.calc_remaind_num()

        if remaind_num >= 0:
            self.is_handoverable_flag = True

        else: self.is_handoverable_flag = False

        return self.is_handoverable_flag
    
    def close_save(self):
        self.wb.save(self.__wb_name)
        self.wb.close()
        
        

if __name__ == "__main__":
    dbManager = DBManager()
    inventory = dbManager.create_inventory(inv_id="FTM003")

    print(inventory.id)
    print(inventory.name)
    print(inventory.quantity)
    print(inventory.min_quantity)
